
#!/usr/bin/env python3
"""
data_generation.py

Comprehensive dataset regeneration pipeline for the University Analytics System.

This script regenerates the major synthetic / anonymized datasets created for the project:

1. Anonymized student master datasets
2. Course catalog with realistic course titles
3. Student grades
4. Student transcripts
5. Academic performance fact tables
6. Payment fact tables
7. Sponsorship / scholarship datasets
8. Attendance datasets
9. Academic progression / retake history
10. Date dimension
11. Documentation markdown files

Author: OpenAI for Final Year Project
License: Project-use
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import math
import os
import random
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd

try:
    import openpyxl  # noqa: F401
except Exception:
    openpyxl = None

# Optional OCR / PDF helpers
try:
    import pytesseract  # type: ignore
    import pdf2image  # type: ignore
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

FIRST_NAMES = [
    "Daniel", "Amina", "Brian", "Grace", "Joseph", "Esther", "Paul", "Ruth",
    "Michael", "Sarah", "Peter", "Joy", "David", "Mary", "Samuel", "Lilian",
    "Isaac", "Patricia", "John", "Hannah", "Noah", "Faith", "Stephen",
    "Dorothy", "Caleb", "Agnes", "Andrew", "Prisca", "Martin", "Stella",
    "Timothy", "Beatrice", "James", "Catherine", "Benjamin", "Deborah",
]
LAST_NAMES = [
    "Okello", "Namakula", "Kato", "Achieng", "Mukasa", "Wekesa", "Akol",
    "Atwine", "Kasule", "Mugisha", "Kagimu", "Namirembe", "Nandutu",
    "Ssebagala", "Muwonge", "Nabirye", "Kawuma", "Nasuuna", "Nanyonjo",
    "Bukenya", "Ssempijja", "Nansubuga", "Kyaligonza", "Tumwebaze",
    "Katusiime", "Bamwine", "Asiimwe", "Nkurunziza", "Mutesi", "Niyonzima",
    "Mutebi", "Ouma",
]

SCH_TYPES = [
    "Merit Award",
    "Need-Based Grant",
    "Leadership Grant",
    "Sports Bursary",
    "Academic Excellence Grant",
    "Innovation Scholarship",
    "Community Grant",
    "Pastoral Scholarship",
]
SCH_NAMES = [
    "Future Leaders Grant",
    "Rising Stars Award",
    "Bright Minds Scholarship",
    "Community Impact Grant",
    "STEM Excellence Award",
    "Global Vision Scholarship",
    "Hope Fund Scholarship",
    "Integrity Award",
    "Service Excellence Award",
    "Pioneer Scholarship",
]

MEX_REASONS = [
    "Financial hold",
    "Medical",
    "No-show",
    "Late registration",
    "Disciplinary",
    "Timetable clash",
    "Other",
]

PAYMENT_METHODS = ["BANK", "MOBILE_MONEY", "CARD", "CASH"]
PAYMENT_CHANNELS = ["BANK_BRANCH", "ONLINE_PORTAL", "MOBILE_APP", "FINANCE_OFFICE"]
PAYMENT_SOURCES = ["PARENT", "SPONSOR", "STUDENT"]
PAYMENT_BANKS = ["Stanbic", "DFCU", "Centenary", "ABSA", "Equity"]
PAYMENT_MOBILES = ["MTN_MOMO", "AIRTEL_MONEY"]

SPONSORS = [
    "Government Sponsorship",
    "Private Sponsor",
    "NGO Education Fund",
    "Church Education Fund",
    "Corporate Scholarship",
    "Alumni Foundation",
    "International Education Grant",
]
SPONSORSHIP_TYPES = [
    "FULL_SCHOLARSHIP",
    "PARTIAL_SCHOLARSHIP",
    "MERIT_BASED_AID",
    "NEED_BASED_AID",
    "TUITION_WAIVER",
]

# Fee rules
UG_FUNCTIONAL = 983_000
PG_FUNCTIONAL = 618_000

HALL_FEES = {
    "EXECUTIVE": {"MUKONO": 1_500_000, "KAMPALA": 1_800_000},
    "DOUBLE_LARGE": {"MUKONO": 1_500_000, "KAMPALA": 1_700_000},
    "DOUBLE_NORMAL": {"MUKONO": 850_000, "KAMPALA": 1_000_000},
    "ORDINARY": {"MUKONO": 650_000, "KAMPALA": 850_000},
}

REG_RE = re.compile(r"^(K)?([JMS])(\d{2})([A-Z])(\d{2})/(\d{3})$", re.I)

# Fallback tuition by broad family when OCR / exact match is unavailable
FALLBACK_TUITION = {
    "IT": 2_425_000,
    "DATA": 2_425_000,
    "BUSINESS": 1_818_000,
    "LAW": 2_325_000,
    "THEOLOGY": 1_818_000,
    "EDUCATION": 1_818_000,
    "HEALTH": 2_586_500,
    "AGRIC": 2_425_000,
    "SOCIAL": 1_818_000,
    "ENGINEERING": 2_545_000,
    "GENERIC_UG": 1_818_000,
    "GENERIC_PG": 2_546_500,
}

COURSE_TITLE_POOLS = {
    "IT": [
        "Introduction to Programming", "Computer Architecture", "Data Structures and Algorithms",
        "Database Systems", "Operating Systems", "Computer Networks", "Web Application Development",
        "Object-Oriented Programming", "Software Engineering", "Systems Analysis and Design",
        "Artificial Intelligence", "Machine Learning Fundamentals", "Cybersecurity Principles",
        "Cloud Computing", "Mobile Application Development", "Human Computer Interaction",
        "Data Mining", "Big Data Analytics", "Information Systems Audit",
        "Research Methods in Computing", "Project Management in IT", "Computer Ethics and Society",
        "Internship and Field Attachment", "Final Year Project",
    ],
    "DATA": [
        "Introduction to Data Science", "Programming for Analytics", "Statistical Methods",
        "Probability and Distributions", "Database Systems", "Data Visualization",
        "Machine Learning I", "Machine Learning II", "Data Warehousing", "Big Data Technologies",
        "Predictive Analytics", "Business Intelligence", "Research Methods", "Ethics in Data Science",
        "Time Series Analysis", "Optimization Methods", "Capstone Project",
    ],
    "BUSINESS": [
        "Principles of Management", "Principles of Accounting", "Business Mathematics",
        "Microeconomics", "Macroeconomics", "Business Communication", "Marketing Principles",
        "Financial Accounting", "Managerial Accounting", "Business Statistics",
        "Organizational Behavior", "Human Resource Management", "Financial Management",
        "Operations Management", "Strategic Management", "Entrepreneurship", "Business Law",
        "Corporate Governance", "Taxation", "Procurement Management", "Supply Chain Management",
        "Internship", "Research Methods", "Project",
    ],
    "LAW": [
        "Legal Methods", "Constitutional Law", "Law of Contract", "Criminal Law",
        "Law of Torts", "Administrative Law", "Property Law", "Commercial Law",
        "Company Law", "Jurisprudence", "International Law", "Human Rights Law",
        "Civil Procedure", "Criminal Procedure", "Evidence", "Family Law",
        "Land Law", "Labour Law", "Alternative Dispute Resolution", "Legal Drafting",
        "Moot Court", "Research Project",
    ],
    "THEOLOGY": [
        "Biblical Studies", "Systematic Theology", "Church History", "Christian Ethics",
        "Homiletics", "Pastoral Theology", "Old Testament Survey", "New Testament Survey",
        "Missiology", "Leadership in Ministry", "Spiritual Formation", "Biblical Hebrew",
        "Biblical Greek", "Hermeneutics", "Counselling in Ministry", "Worship and Liturgy",
        "Research Methods in Theology", "Field Education", "Dissertation",
    ],
    "EDUCATION": [
        "Foundations of Education", "Educational Psychology", "Curriculum Studies",
        "Classroom Management", "Assessment and Evaluation", "Instructional Methods",
        "Guidance and Counselling", "Education Administration", "Sociology of Education",
        "Philosophy of Education", "Educational Research Methods", "Special Needs Education",
        "Teaching Practice", "Measurement and Evaluation", "ICT in Education",
        "Comparative Education", "Project",
    ],
    "HEALTH": [
        "Human Anatomy", "Human Physiology", "Biochemistry", "Pharmacology",
        "Medical Microbiology", "Pathology", "Community Health", "Principles of Nursing",
        "Adult Health Nursing", "Maternal and Child Health", "Mental Health Nursing",
        "Nutrition", "Epidemiology", "Research Methods in Health Sciences",
        "Clinical Practice", "Health Promotion", "Public Health Policy", "Capstone Project",
    ],
    "AGRIC": [
        "Introduction to Agriculture", "Soil Science", "Crop Science", "Animal Science",
        "Agricultural Economics", "Agribusiness Management", "Agricultural Extension",
        "Plant Pathology", "Farm Management", "Irrigation and Water Management",
        "Agricultural Marketing", "Entrepreneurship in Agriculture", "Statistics for Agriculture",
        "Research Methods", "Field Attachment", "Project",
    ],
    "SOCIAL": [
        "Introduction to Sociology", "Social Work Practice", "Community Development",
        "Counselling Skills", "Social Policy", "Research Methods", "Development Studies",
        "Governance and Public Policy", "International Relations Theory", "Conflict Resolution",
        "Public Administration", "Gender Studies", "Humanitarian Studies", "Field Work", "Project",
    ],
    "ENGINEERING": [
        "Engineering Mathematics", "Engineering Drawing", "Mechanics", "Materials Science",
        "Fluid Mechanics", "Thermodynamics", "Electrical Principles", "Surveying",
        "Water and Sanitation Systems", "Environmental Engineering", "Structural Analysis",
        "Project Planning", "GIS and Remote Sensing", "Field Attachment", "Design Project",
    ],
    "GENERIC": [
        "Communication Skills", "Critical Thinking", "Research Methods", "Entrepreneurship",
        "Ethics and Integrity", "Statistics", "Project Planning", "Field Attachment",
        "Internship", "Capstone Project", "Professional Practice", "Seminar",
    ],
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def stable_hash(text: str, salt: str = "UCU_DEMO_SALT_2026") -> int:
    h = hashlib.sha256((salt + "|" + str(text)).encode("utf-8")).hexdigest()
    return int(h[:16], 16)


def stable_pick(items: List[str], key: str) -> str:
    return items[stable_hash(key) % len(items)]


def fake_person_name(key: str) -> str:
    return f"{stable_pick(FIRST_NAMES, 'F|' + key)} {stable_pick(LAST_NAMES, 'L|' + key)}"


def rand_ref(length: int = 10) -> str:
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    return "".join(random.choice(chars) for _ in range(length))


def find_header_row(excel_path: Path, sheet_name: str = "Main") -> int:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to detect header rows in the source workbooks.")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    for r in range(1, 80):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "S/N" in vals and "PROGRAM" in vals and "RETAKE COURSES" in vals:
            return r
    raise ValueError(f"Could not locate header row in {excel_path}")


def load_source_students(path: Path) -> pd.DataFrame:
    hdr = find_header_row(path)
    return pd.read_excel(path, sheet_name="Main", header=hdr - 1)


def parse_reg(regno: str) -> Optional[Dict[str, object]]:
    if not isinstance(regno, str):
        return None
    m = REG_RE.match(regno.strip().upper())
    if not m:
        return None
    k, intake, yy, level, code, serial = m.groups()
    return {
        "kampala": bool(k),
        "intake": intake,
        "yy": int(yy),
        "level": level,
        "code": int(code),
        "serial": int(serial),
    }


def academic_year_for_reg(regno: str, semester_index: int) -> Tuple[str, str]:
    parts = parse_reg(regno)
    start_year = 2024 if not parts else 2000 + int(parts["yy"])
    y = start_year + (semester_index - 1) // 2
    semester = "SEM1" if semester_index % 2 == 1 else "SEM2"
    return f"{y}/{y+1}", semester


def campus_from_reg(regno: str) -> str:
    parts = parse_reg(regno)
    if not parts:
        return "MUKONO"
    return "KAMPALA" if parts["kampala"] else "MUKONO"


def is_undergrad(regno: str, program: str = "") -> bool:
    parts = parse_reg(regno)
    if parts:
        lvl = str(parts["level"]).upper()
        if lvl in {"M", "P"}:
            return False
        if lvl in {"B", "D", "C", "H"}:
            return True
    p = str(program).lower()
    return not p.startswith(("master", "phd", "doctor", "pgd"))


def norm_accom_type(val: object) -> Optional[str]:
    if not isinstance(val, str):
        return None
    u = val.strip().upper()
    if "EXEC" in u:
        return "EXECUTIVE"
    if "DOUBLE" in u and ("LARGE" in u or "BIG" in u):
        return "DOUBLE_LARGE"
    if "DOUBLE" in u:
        return "DOUBLE_NORMAL"
    if "ORD" in u:
        return "ORDINARY"
    return None


def parse_pct(value: object) -> float:
    if pd.isna(value):
        return 0.0
    m = re.search(r"(\d+(\.\d+)?)", str(value))
    return float(m.group(1)) if m else 0.0


def pct_bucket(p: float) -> str:
    if p <= 25:
        return "0–25%"
    if p <= 50:
        return "26–50%"
    if p <= 75:
        return "51–75%"
    return "76–100%"


def clean_retake_value(v: object) -> str:
    if pd.isna(v):
        return ""
    s = str(v).strip()
    if s == "" or s.upper() in {"NONE", "N/A", "NA"}:
        return ""
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"\s*;\s*", ", ", s)
    s = re.sub(r"\s{2,}", " ", s)
    return s


def program_family(program: str) -> str:
    p = str(program).lower()
    if any(k in p for k in ["information technology", "computer", "computing", "software", "information systems", "cyber"]):
        return "IT"
    if any(k in p for k in ["data science", "analytics", "statistics"]):
        return "DATA"
    if any(k in p for k in ["business", "accounting", "finance", "economics", "procurement", "human resource", "tourism", "hospitality", "management", "agribusiness"]):
        return "BUSINESS"
    if "law" in p:
        return "LAW"
    if any(k in p for k in ["divinity", "theology", "ministry", "biblical"]):
        return "THEOLOGY"
    if any(k in p for k in ["education", "curriculum", "pedagogy", "library and information science"]):
        return "EDUCATION"
    if any(k in p for k in ["nursing", "public health", "health", "medicine", "clinical", "nutrition"]):
        return "HEALTH"
    if any(k in p for k in ["agric", "food science", "crop", "animal"]):
        return "AGRIC"
    if any(k in p for k in ["social work", "social administration", "governance", "international relations", "development", "public administration", "community"]):
        return "SOCIAL"
    if any(k in p for k in ["engineering", "water and sanitation", "design", "technology"]):
        return "ENGINEERING"
    return "GENERIC"


def program_prefix(program: str) -> str:
    words = [w for w in re.split(r"[^A-Za-z]+", str(program).upper()) if w]
    if len(words) == 1:
        return (words[0][:3] + "XXX")[:3]
    return ("".join(w[0] for w in words[:3]) + "XXX")[:3]


# ---------------------------------------------------------------------------
# Fee extraction
# ---------------------------------------------------------------------------

def extract_tuition_map_from_pdf(pdf_path: Path) -> Dict[str, int]:
    """
    Attempts OCR extraction of tuition values from the fee structure PDF.
    Returns a map of program_name -> nationals tuition.
    If OCR libraries are unavailable or extraction fails, returns {}.
    """
    if not OCR_AVAILABLE or not pdf_path.exists():
        return {}

    tuition_map: Dict[str, int] = {}
    try:
        images = pdf2image.convert_from_path(str(pdf_path), dpi=180)
        prog_re = re.compile(r"^(Diploma|Bachelor|Master|PhD|Doctor|PGD|Higher Education Certificate)", re.I)
        for img in images:
            text = pytesseract.image_to_string(img)
            for line in text.splitlines():
                line = line.strip()
                if not line or not prog_re.match(line):
                    continue
                nums = re.findall(r"(\d{1,3}(?:,\d{3})+)", line)
                if not nums:
                    continue
                prog = re.sub(r"(\d{1,3}(?:,\d{3})+)", "", line).strip()
                prog = re.sub(r"\s{2,}", " ", prog).replace("|", "").replace(" il", "").strip()
                prog = re.sub(r"\s+0$", "", prog)
                tuition_map[prog] = int(nums[0].replace(",", ""))
    except Exception:
        return {}
    return tuition_map


def tuition_for_program(program: str, regno: str, tuition_map: Dict[str, int]) -> int:
    if not isinstance(program, str) or not program.strip():
        return 0
    p = program.strip()
    if p in tuition_map:
        return tuition_map[p]
    # fuzzy fallback without external dependency
    best_name = None
    best_score = 0.0
    for cand in tuition_map.keys():
        score = _simple_similarity(p.lower(), cand.lower())
        if score > best_score:
            best_score = score
            best_name = cand
    if best_name and best_score >= 0.86:
        return tuition_map[best_name]

    family = program_family(p)
    if family == "GENERIC":
        return FALLBACK_TUITION["GENERIC_UG"] if is_undergrad(regno, p) else FALLBACK_TUITION["GENERIC_PG"]
    return FALLBACK_TUITION[family]


def _simple_similarity(a: str, b: str) -> float:
    # lightweight similarity: Jaccard on tokens + prefix bonus
    ta = set(a.split())
    tb = set(b.split())
    inter = len(ta & tb)
    union = max(1, len(ta | tb))
    base = inter / union
    if a[:12] == b[:12]:
        base += 0.2
    return min(base, 1.0)


# ---------------------------------------------------------------------------
# Students generation
# ---------------------------------------------------------------------------

def build_program_signature(df: pd.DataFrame) -> Dict[str, Tuple[str, str]]:
    sig: Dict[str, List[Tuple[str, str]]] = {}
    for reg, prog in zip(df["REG. NO."], df["PROGRAM"]):
        parts = parse_reg(str(reg))
        if not parts or not isinstance(prog, str) or not prog.strip():
            continue
        sig.setdefault(prog.strip(), []).append((str(parts["level"]), f'{int(parts["code"]):02d}'))
    out: Dict[str, Tuple[str, str]] = {}
    for prog, pairs in sig.items():
        counts: Dict[Tuple[str, str], int] = {}
        for pair in pairs:
            counts[pair] = counts.get(pair, 0) + 1
        out[prog] = max(counts.items(), key=lambda kv: kv[1])[0]
    return out


def build_intake_year_dist(df: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, int]]]:
    dist: Dict[str, Dict[str, Dict[str, int]]] = {}
    for reg, prog in zip(df["REG. NO."], df["PROGRAM"]):
        parts = parse_reg(str(reg))
        if not parts or not isinstance(prog, str) or not prog.strip():
            continue
        d = dist.setdefault(prog.strip(), {"intake": {}, "yy": {}})
        intake = str(parts["intake"])
        yy = f'{int(parts["yy"]):02d}'
        d["intake"][intake] = d["intake"].get(intake, 0) + 1
        d["yy"][yy] = d["yy"].get(yy, 0) + 1
    return dist


def deterministic_counter_pick(counter: Dict[str, int], key: str, fallback: List[str]) -> str:
    if not counter:
        return stable_pick(fallback, key)
    items = list(counter.items())
    labels = [k for k, _ in items]
    weights = np.array([v for _, v in items], dtype=float)
    weights = weights / weights.sum()
    r = (stable_hash(key) % 10_000) / 10_000.0
    cum = np.cumsum(weights)
    idx = int(np.searchsorted(cum, r, side="right"))
    idx = min(idx, len(labels) - 1)
    return labels[idx]


def generate_unique_reg(
    program: str,
    used: set,
    serial_counters: Dict[Tuple[str, str, str, str], int],
    program_signature: Dict[str, Tuple[str, str]],
    intake_year_dist: Dict[str, Dict[str, Dict[str, int]]],
    campus_prefix: Optional[str] = None,
) -> str:
    lvl, code = program_signature.get(program, ("B", "99"))
    d = intake_year_dist.get(program, {"intake": {}, "yy": {}})
    intake = deterministic_counter_pick(d.get("intake", {}), f"INT|{program}|{len(used)}", ["J", "M", "S"])
    yy = deterministic_counter_pick(d.get("yy", {}), f"YY|{program}|{len(used)}", ["23", "24", "25"])
    prefix = "K" if campus_prefix == "KAMPALA" else ""
    key = (prefix + intake, yy, lvl, code)
    serial_counters[key] = serial_counters.get(key, 0) + 1
    while True:
        reg = f"{prefix}{intake}{yy}{lvl}{code}/{serial_counters[key]:03d}"
        if reg not in used:
            used.add(reg)
            return reg
        serial_counters[key] += 1


def generate_unique_acc(existing: set, count: int) -> List[str]:
    used = set(existing)
    out = []
    for i in range(count):
        base = "B"
        n = 10000 + i
        candidate = f"{base}{n}"
        while candidate in used:
            n += 1
            candidate = f"{base}{n}"
        used.add(candidate)
        out.append(candidate)
    return out


def build_retake_pool_by_program(df: pd.DataFrame) -> Tuple[Dict[str, List[str]], Dict[str, float], List[str], float]:
    pools: Dict[str, List[str]] = {}
    none_rate: Dict[str, float] = {}
    for prog, grp in df.groupby(df["PROGRAM"].astype(str).str.strip()):
        if not prog or prog.lower() == "nan":
            continue
        vals = grp["RETAKE COURSES"].apply(clean_retake_value)
        non_empty = sorted(set([x for x in vals.tolist() if x != ""]))
        pools[prog] = non_empty
        none_rate[prog] = float((vals == "").mean())
    all_vals = df["RETAKE COURSES"].apply(clean_retake_value)
    global_pool = sorted(set([x for x in all_vals.tolist() if x != ""]))
    global_none = float((all_vals == "").mean())
    return pools, none_rate, global_pool, global_none


def assign_program_linked_retakes(
    df: pd.DataFrame,
    pools: Dict[str, List[str]],
    none_rate: Dict[str, float],
    global_pool: List[str],
    global_none: float,
    tag: str,
) -> pd.Series:
    out = []
    for i, (prog, reg) in enumerate(zip(df["PROGRAM"].astype(str).str.strip(), df["REG. NO."].astype(str).str.strip())):
        key = f"{tag}|{prog}|{reg}|{i}"
        pool = pools.get(prog, [])
        pr_none = none_rate.get(prog, global_none)
        u = (stable_hash("NONE|" + key, salt="RET_SALT_2026") % 10_000) / 10_000.0
        if u < pr_none or (not pool and not global_pool):
            out.append("")
            continue
        use_pool = pool if pool else global_pool
        out.append(use_pool[stable_hash("PICK|" + key, salt="RET_SALT_2026") % len(use_pool)])
    return pd.Series(out)


def anonymize_students(
    df: pd.DataFrame,
    source_tag: str,
    target_rows: int,
    tuition_map: Dict[str, int],
    retake_source: pd.DataFrame,
    seed: int = 20260304,
) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    out = df.copy()

    # Consistent fake names for students and officers
    out["NAME"] = out["NAME"].apply(lambda x: fake_person_name(f"STUDENT|{source_tag}|{x}") if isinstance(x, str) and x.strip() else x)

    for col, prefix in [
        ("ACADEMIC OFFICER", "AO"),
        ("ACCOMMODATION OFFICER", "ACCO"),
        ("ACCOUNTS OFFICER", "ACCTS"),
        ("FINANCIAL AID OFFICER", "FA"),
        ("MEALS OFFICER", "MEALS"),
    ]:
        if col in out.columns:
            unique_vals = [u for u in out[col].dropna().astype(str).str.strip().unique() if u and u.upper() != "NONE" and u.lower() != "nan"]
            mapping = {u: fake_person_name(f"{source_tag}|{prefix}|{u}") for u in sorted(unique_vals)}
            out[col] = out[col].apply(lambda v: mapping.get(str(v).strip(), v) if pd.notna(v) else v)

    if "SCHOLARSHIP TYPE" in out.columns:
        unique_vals = [u for u in out["SCHOLARSHIP TYPE"].dropna().astype(str).str.strip().unique() if u and u.upper() != "NONE" and u.lower() != "nan"]
        mapping = {u: SCH_TYPES[i % len(SCH_TYPES)] for i, u in enumerate(sorted(unique_vals))}
        out["SCHOLARSHIP TYPE"] = out["SCHOLARSHIP TYPE"].apply(lambda v: mapping.get(str(v).strip(), v) if pd.notna(v) else v)

    if "SCHOLARSHIP NAME" in out.columns:
        unique_vals = [u for u in out["SCHOLARSHIP NAME"].dropna().astype(str).str.strip().unique() if u and u.upper() != "NONE" and u.lower() != "nan"]
        mapping = {u: SCH_NAMES[i % len(SCH_NAMES)] for i, u in enumerate(sorted(unique_vals))}
        out["SCHOLARSHIP NAME"] = out["SCHOLARSHIP NAME"].apply(lambda v: mapping.get(str(v).strip(), v) if pd.notna(v) else v)

    # Keep original existing REG NO / ACC NO where present; expand by cloning
    used_reg = set(out["REG. NO."].dropna().astype(str).str.strip().str.upper())
    used_acc = set(out["ACC. NO."].dropna().astype(str).str.strip().str.upper())
    all_source = pd.concat([df, retake_source], ignore_index=True)
    prog_sig = build_program_signature(all_source)
    intake_dist = build_intake_year_dist(all_source)

    serial_counters: Dict[Tuple[str, str, str, str], int] = {}
    for reg in used_reg:
        parts = parse_reg(reg)
        if parts:
            key = (
                ("K" if parts["kampala"] else "") + str(parts["intake"]),
                f'{int(parts["yy"]):02d}',
                str(parts["level"]),
                f'{int(parts["code"]):02d}',
            )
            serial_counters[key] = max(serial_counters.get(key, 0), int(parts["serial"]))

    if len(out) < target_rows:
        need = target_rows - len(out)
        sample_idx = rng.choice(out.index.to_numpy(), size=need, replace=True)
        extra = out.loc[sample_idx].copy().reset_index(drop=True)

        # New access numbers
        extra["ACC. NO."] = generate_unique_acc(used_acc, need)

        # New REG NO values, preserving program linkage + campus prefix if existing record has Kampala pattern
        generated_regs = []
        for _, row in extra.iterrows():
            prog = str(row["PROGRAM"]).strip()
            current_reg = str(row["REG. NO."]).strip()
            campus_pref = campus_from_reg(current_reg)
            generated_regs.append(generate_unique_reg(prog, used_reg, serial_counters, prog_sig, intake_dist, campus_pref))
        extra["REG. NO."] = generated_regs

        # New fake names keyed to new reg no
        extra["NAME"] = extra["REG. NO."].apply(lambda r: fake_person_name(f"STUDENT|{source_tag}|{r}"))
        extra["S/N"] = np.arange(len(out) + 1, target_rows + 1)

        out["S/N"] = np.arange(1, len(out) + 1)
        out = pd.concat([out, extra], ignore_index=True)

    # Fee logic
    pct = out["FEES"].apply(parse_pct)
    out["EXPECTED FEES(%) - RAW"] = pct
    out["EXPECTED FEES(%) - RANGE"] = pct.apply(pct_bucket)

    tuition = []
    functional = []
    hall_fee = []
    for reg, prog, res, accom in zip(out["REG. NO."], out["PROGRAM"], out["RESIDENCE"], out["ACCOMMODATION Type"]):
        reg = str(reg)
        prog = str(prog)
        tuition.append(tuition_for_program(prog, reg, tuition_map))
        functional.append(0 if "(blended)" in prog.lower() else (UG_FUNCTIONAL if is_undergrad(reg, prog) else PG_FUNCTIONAL))
        hf = 0
        if str(res).strip().upper() == "R":
            cat = norm_accom_type(accom)
            if cat is not None:
                hf = HALL_FEES[cat][campus_from_reg(reg)]
        hall_fee.append(hf)

    out["TOTAL EXPECTED FEES (UGX)"] = (np.array(tuition) + np.array(functional) + np.array(hall_fee)).astype(int)
    out["AMOUNT PAID (UGX)"] = np.rint(out["TOTAL EXPECTED FEES (UGX)"] * (pct / 100.0)).astype(int)
    out["TUITION BALANCE (UGX)"] = (out["TOTAL EXPECTED FEES (UGX)"] - out["AMOUNT PAID (UGX)"]).astype(int)
    out.loc[out["TUITION BALANCE (UGX)"] < 0, "TUITION BALANCE (UGX)"] = 0

    # Program-linked retakes using ONLY original source values
    pools, none_rate, global_pool, global_none = build_retake_pool_by_program(retake_source)
    out["RETAKE COURSES"] = assign_program_linked_retakes(out, pools, none_rate, global_pool, global_none, source_tag)

    return out


# ---------------------------------------------------------------------------
# Course catalog
# ---------------------------------------------------------------------------

def titles_for_program(program: str, n_needed: int) -> List[str]:
    family = program_family(program)
    pool = COURSE_TITLE_POOLS.get(family, COURSE_TITLE_POOLS["GENERIC"])
    titles: List[str] = []
    idx = 0
    while len(titles) < n_needed:
        title = pool[idx % len(pool)]
        if idx >= len(pool):
            cycle = idx // len(pool) + 1
            title = f"{title} {cycle}"
        titles.append(title)
        idx += 1
    return titles


def semester_course_count(program: str, sem: int) -> int:
    return 6 + ((sum(ord(c) for c in program) + sem) % 2)


def build_course_catalog(programs: Iterable[str], max_semesters: int) -> pd.DataFrame:
    rows = []
    for prog in sorted(set(programs)):
        prefix = program_prefix(prog)
        total_needed = sum(semester_course_count(prog, s) for s in range(1, max_semesters + 1))
        titles = titles_for_program(prog, total_needed)
        cursor = 0
        for sem in range(1, max_semesters + 1):
            n = semester_course_count(prog, sem)
            for i in range(1, n + 1):
                rows.append({
                    "PROGRAM": prog,
                    "SEMESTER_INDEX": sem,
                    "COURSE_CODE": f"{prefix}{sem:02d}{i:02d}",
                    "COURSE_TITLE": titles[cursor],
                    "COURSE_UNITS": [2, 3, 3, 3, 4][(sem + i) % 5],
                    "COURSE_TYPE": "CORE" if i <= max(4, n - 1) else "ELECTIVE",
                })
                cursor += 1
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Grades
# ---------------------------------------------------------------------------

def course_result(rng: np.random.Generator, base_ability: float) -> Tuple[float, float, str, str, float]:
    u = float(rng.random())
    if u < 0.05:
        cw = int(rng.integers(0, 17))
        return float(cw), float("nan"), "FCW", "", float(cw)
    if u < 0.08:
        cw = int(np.clip(rng.normal(32 + 18 * base_ability, 8), 17, 60))
        return float(cw), float("nan"), "MEX", rng.choice(MEX_REASONS), float("nan")

    cw = int(np.clip(rng.normal(30 + 22 * base_ability, 9), 0, 60))
    if cw < 17:
        cw = int(rng.integers(0, 17))
        return float(cw), float("nan"), "FCW", "", float(cw)

    exam = int(np.clip(rng.normal(18 + 18 * base_ability, 7), 0, 40))
    final = float(cw + exam)
    if exam < 17 or final < 50:
        return float(cw), float(exam), "FEX", "", final
    return float(cw), float(exam), "Completed", "", final


def letter_grade_and_points(final_mark: float, status: str) -> Tuple[str, float]:
    if status != "Completed":
        return "F", 1.5
    m = float(final_mark)
    if m >= 80: return "A", 5.0
    if m >= 75: return "B+", 4.5
    if m >= 70: return "B", 4.0
    if m >= 65: return "C+", 3.5
    if m >= 60: return "C", 3.0
    if m >= 55: return "D+", 2.5
    if m >= 50: return "D", 2.0
    return "F", 1.5


def build_catalog_lookup(catalog: pd.DataFrame) -> Dict[Tuple[str, int], List[Tuple[str, str, int]]]:
    lookup: Dict[Tuple[str, int], List[Tuple[str, str, int]]] = {}
    for (prog, sem), grp in catalog.groupby(["PROGRAM", "SEMESTER_INDEX"], sort=False):
        lookup[(prog, int(sem))] = list(grp[["COURSE_CODE", "COURSE_TITLE", "COURSE_UNITS"]].itertuples(index=False, name=None))
    return lookup


def generate_grades(students: pd.DataFrame, catalog: pd.DataFrame, tag: str, seed: int) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    catalog_lookup = build_catalog_lookup(catalog)
    rows = []
    for idx, stu in students.iterrows():
        regno = str(stu["REG. NO."]).strip()
        accno = str(stu.get("ACC. NO.", "")).strip()
        program = str(stu.get("PROGRAM", "")).strip()
        total_regs = int(stu.get("TOTAL REGISTRATIONS", 1) if not pd.isna(stu.get("TOTAL REGISTRATIONS", 1)) else 1)
        fees_pct = float(stu.get("EXPECTED FEES(%) - RAW", 0) if not pd.isna(stu.get("EXPECTED FEES(%) - RAW", 0)) else 0.0)
        base_ability = float(np.clip(rng.normal(0.45 + 0.002 * (fees_pct - 50), 0.18), 0.05, 0.95))

        for sem in range(1, total_regs + 1):
            ay, sem_label = academic_year_for_reg(regno, sem)
            courses = catalog_lookup.get((program, sem), catalog_lookup.get((program, 1), []))
            for course_code, course_title, course_units in courses:
                cw, exam, status, mex_reason, final = course_result(rng, base_ability)
                letter, gp = letter_grade_and_points(0 if math.isnan(final) else final, status)
                if status == "MEX":
                    letter, gp = "F", 1.5
                rows.append({
                    "RECORD_ID": f"{tag}-GR-{idx+1:05d}-{sem:02d}-{course_code}",
                    "REG_NO": regno,
                    "ACC_NO": accno,
                    "PROGRAM": program,
                    "ACADEMIC_YEAR": ay,
                    "SEMESTER": sem_label,
                    "SEMESTER_INDEX": sem,
                    "COURSE_CODE": course_code,
                    "COURSE_TITLE": course_title,
                    "COURSE_UNITS": int(course_units),
                    "CW_MARK_60": cw,
                    "EXAM_MARK_40": exam,
                    "FINAL_MARK_100": final,
                    "STATUS": status,
                    "MEX_REASON": mex_reason,
                    "LETTER_GRADE": letter,
                    "GRADE_POINTS": gp,
                })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Transcripts and performance
# ---------------------------------------------------------------------------

def build_transcript(grades: pd.DataFrame) -> pd.DataFrame:
    x = grades.copy()
    x["COURSE_UNITS"] = pd.to_numeric(x["COURSE_UNITS"], errors="coerce").fillna(0.0)
    x["GRADE_POINTS"] = pd.to_numeric(x["GRADE_POINTS"], errors="coerce").fillna(0.0)
    x["FINAL_MARK_100"] = pd.to_numeric(x["FINAL_MARK_100"], errors="coerce")
    x["QUALITY_POINTS_ROW"] = x["COURSE_UNITS"] * x["GRADE_POINTS"]
    x["PASSED_CREDITS_ROW"] = np.where((x["STATUS"] == "Completed") & (x["FINAL_MARK_100"] >= 50), x["COURSE_UNITS"], 0.0)
    x["IS_COMPLETED"] = (x["STATUS"] == "Completed").astype(int)
    x["IS_FAILED"] = (x["STATUS"] != "Completed").astype(int)
    x["IS_FCW"] = (x["STATUS"] == "FCW").astype(int)
    x["IS_FEX"] = (x["STATUS"] == "FEX").astype(int)
    x["IS_MEX"] = (x["STATUS"] == "MEX").astype(int)

    group_cols = ["REG_NO", "ACC_NO", "PROGRAM", "ACADEMIC_YEAR", "SEMESTER", "SEMESTER_INDEX"]
    sem = x.groupby(group_cols, as_index=False)[
        ["COURSE_UNITS", "QUALITY_POINTS_ROW", "COURSE_CODE", "IS_COMPLETED", "IS_FAILED", "IS_FCW", "IS_FEX", "IS_MEX", "PASSED_CREDITS_ROW"]
    ].agg({
        "COURSE_UNITS": "sum",
        "QUALITY_POINTS_ROW": "sum",
        "COURSE_CODE": "count",
        "IS_COMPLETED": "sum",
        "IS_FAILED": "sum",
        "IS_FCW": "sum",
        "IS_FEX": "sum",
        "IS_MEX": "sum",
        "PASSED_CREDITS_ROW": "sum",
    })

    sem = sem.rename(columns={
        "COURSE_UNITS": "CREDITS_ATTEMPTED",
        "QUALITY_POINTS_ROW": "QUALITY_POINTS",
        "COURSE_CODE": "COURSES_COUNT",
        "IS_COMPLETED": "PASSED_COURSES",
        "IS_FAILED": "FAILED_COURSES",
        "IS_FCW": "FCW_COUNT",
        "IS_FEX": "FEX_COUNT",
        "IS_MEX": "MEX_COUNT",
        "PASSED_CREDITS_ROW": "CREDITS_PASSED",
    })

    sem["CREDITS_FAILED"] = sem["CREDITS_ATTEMPTED"] - sem["CREDITS_PASSED"]
    sem["SEMESTER_GPA"] = np.where(sem["CREDITS_ATTEMPTED"] > 0, sem["QUALITY_POINTS"] / sem["CREDITS_ATTEMPTED"], np.nan)

    sem = sem.sort_values(["REG_NO", "SEMESTER_INDEX"]).reset_index(drop=True)
    sem["CUM_CREDITS_ATTEMPTED"] = sem.groupby("REG_NO")["CREDITS_ATTEMPTED"].cumsum()
    sem["CUM_CREDITS_PASSED"] = sem.groupby("REG_NO")["CREDITS_PASSED"].cumsum()
    sem["CUM_QUALITY_POINTS"] = sem.groupby("REG_NO")["QUALITY_POINTS"].cumsum()
    sem["CGPA"] = np.where(sem["CUM_CREDITS_ATTEMPTED"] > 0, sem["CUM_QUALITY_POINTS"] / sem["CUM_CREDITS_ATTEMPTED"], np.nan)

    round_cols = [
        "CREDITS_ATTEMPTED", "QUALITY_POINTS", "CREDITS_PASSED", "CREDITS_FAILED",
        "SEMESTER_GPA", "CUM_CREDITS_ATTEMPTED", "CUM_CREDITS_PASSED", "CUM_QUALITY_POINTS", "CGPA",
    ]
    sem[round_cols] = sem[round_cols].round(2)
    return sem


def build_academic_performance(grades: pd.DataFrame) -> pd.DataFrame:
    g = grades.copy()
    g["COURSE_UNITS"] = pd.to_numeric(g["COURSE_UNITS"], errors="coerce").fillna(0)
    g["GRADE_POINTS"] = pd.to_numeric(g["GRADE_POINTS"], errors="coerce").fillna(0)
    g["FINAL_MARK_100"] = pd.to_numeric(g["FINAL_MARK_100"], errors="coerce")
    g["QUALITY_POINTS"] = g["COURSE_UNITS"] * g["GRADE_POINTS"]
    g["PASSED"] = ((g["STATUS"] == "Completed") & (g["FINAL_MARK_100"] >= 50)).astype(int)
    g["FAILED"] = (g["STATUS"] != "Completed").astype(int)
    g["FCW"] = (g["STATUS"] == "FCW").astype(int)
    g["FEX"] = (g["STATUS"] == "FEX").astype(int)
    g["MEX"] = (g["STATUS"] == "MEX").astype(int)

    group_cols = ["REG_NO", "ACC_NO", "PROGRAM", "ACADEMIC_YEAR", "SEMESTER", "SEMESTER_INDEX"]
    fact = g.groupby(group_cols, as_index=False).agg(
        COURSES_REGISTERED=("COURSE_CODE", "count"),
        TOTAL_CREDITS=("COURSE_UNITS", "sum"),
        QUALITY_POINTS=("QUALITY_POINTS", "sum"),
        PASSED_COURSES=("PASSED", "sum"),
        FAILED_COURSES=("FAILED", "sum"),
        FCW_COUNT=("FCW", "sum"),
        FEX_COUNT=("FEX", "sum"),
        MEX_COUNT=("MEX", "sum"),
    )
    fact["SEMESTER_GPA"] = np.where(fact["TOTAL_CREDITS"] > 0, fact["QUALITY_POINTS"] / fact["TOTAL_CREDITS"], np.nan).round(2)
    return fact


def build_academic_progression(grades: pd.DataFrame, tag: str) -> pd.DataFrame:
    df = grades.sort_values(["REG_NO", "COURSE_CODE", "SEMESTER_INDEX"]).copy()
    df["ATTEMPT_NUMBER"] = df.groupby(["REG_NO", "COURSE_CODE"]).cumcount() + 1
    df["RETAKE_FLAG"] = df["STATUS"].isin({"FEX", "FCW", "MEX"})
    df["PROGRESSION_STATUS"] = np.where(df["RETAKE_FLAG"], "RETAKE_REQUIRED", "NORMAL")
    df["PROGRESSION_ID"] = tag + "-" + df["REG_NO"].astype(str) + "-" + df["COURSE_CODE"].astype(str) + "-" + df["ATTEMPT_NUMBER"].astype(str)
    return df[["PROGRESSION_ID", "REG_NO", "COURSE_CODE", "SEMESTER_INDEX", "STATUS", "ATTEMPT_NUMBER", "RETAKE_FLAG", "PROGRESSION_STATUS"]]


# ---------------------------------------------------------------------------
# Finance
# ---------------------------------------------------------------------------

def generate_payments(students: pd.DataFrame, tag: str, seed: int) -> pd.DataFrame:
    rng = random.Random(seed)
    rows = []
    pay_id = 1
    for _, r in students.iterrows():
        reg = r["REG. NO."]
        acc = r.get("ACC. NO.", "")
        regs = int(r.get("TOTAL REGISTRATIONS", 1) if not pd.isna(r.get("TOTAL REGISTRATIONS", 1)) else 1)
        total = float(r.get("TOTAL EXPECTED FEES (UGX)", 0) or 0)
        incomplete = rng.random() < 0.02

        for sem in range(1, regs + 1):
            ay, sem_label = academic_year_for_reg(str(reg), sem)
            if sem == regs and incomplete:
                target = total * rng.uniform(0.4, 0.8)
            else:
                target = total * rng.uniform(0.95, 1.05)
            n = rng.randint(1, 6)
            remaining = target
            start_year = int(ay.split("/")[0])
            start_month = 1 if sem_label == "SEM1" else 8

            for p in range(n):
                if p == n - 1:
                    amt = max(0.0, remaining)
                else:
                    amt = rng.uniform(0.1, 0.4) * target / n * 3
                    amt = min(amt, remaining)
                remaining -= amt

                method = rng.choice(PAYMENT_METHODS)
                bank = rng.choice(PAYMENT_BANKS) if method == "BANK" else ""
                mobile = rng.choice(PAYMENT_MOBILES) if method == "MOBILE_MONEY" else ""
                payment_date = dt.date(start_year, start_month, 1) + dt.timedelta(days=rng.randint(0, 60))

                rows.append({
                    "PAYMENT_ID": f"{tag}-{pay_id}",
                    "REG_NO": reg,
                    "ACC_NO": acc,
                    "ACADEMIC_YEAR": ay,
                    "SEMESTER": sem_label,
                    "SEMESTER_INDEX": sem,
                    "PAYMENT_DATE": payment_date,
                    "AMOUNT_UGX": int(round(amt)),
                    "PAYMENT_METHOD": method,
                    "BANK_NAME": bank,
                    "MOBILE_PROVIDER": mobile,
                    "CHANNEL": rng.choice(PAYMENT_CHANNELS),
                    "PAYMENT_SOURCE": rng.choice(PAYMENT_SOURCES),
                    "TRANSACTION_REFERENCE": rand_ref(12),
                    "PAYMENT_STATUS": rng.choice(["SUCCESS", "SUCCESS", "SUCCESS", "SUCCESS", "FAILED"]),
                })
                pay_id += 1
    return pd.DataFrame(rows)


def generate_sponsorships(students: pd.DataFrame, tag: str, seed: int) -> pd.DataFrame:
    rng = random.Random(seed)
    rows = []
    pid = 1
    for _, r in students.iterrows():
        reg = r["REG. NO."]
        acc = r.get("ACC. NO.", "")
        regs = int(r.get("TOTAL REGISTRATIONS", 1) if not pd.isna(r.get("TOTAL REGISTRATIONS", 1)) else 1)
        expected = float(r.get("TOTAL EXPECTED FEES (UGX)", 0) or 0)
        if rng.random() > rng.uniform(0.15, 0.25):
            continue
        sponsor = rng.choice(SPONSORS)
        sch_type = rng.choice(SPONSORSHIP_TYPES)
        for sem in range(1, regs + 1):
            ay, _ = academic_year_for_reg(str(reg), sem)
            coverage = rng.uniform(0.2, 1.0)
            amount = int(round(expected * coverage))
            rows.append({
                "SPONSOR_PAYMENT_ID": f"SP-{tag}-{pid}",
                "REG_NO": reg,
                "ACC_NO": acc,
                "SPONSOR_NAME": sponsor,
                "SCHOLARSHIP_TYPE": sch_type,
                "SEMESTER_INDEX": sem,
                "ACADEMIC_YEAR": ay,
                "AMOUNT_SPONSORED_UGX": amount,
                "DISBURSEMENT_DATE": dt.date(rng.randint(2022, 2026), rng.randint(1, 12), rng.randint(1, 28)),
                "SPONSOR_REFERENCE": rand_ref(10),
                "STATUS": rng.choice(["APPROVED", "APPROVED", "APPROVED", "PENDING"]),
            })
            pid += 1
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Attendance and date dimension
# ---------------------------------------------------------------------------

def semester_start(year: int, intake: str) -> dt.date:
    if intake == "J":
        return dt.date(year, 1, 10)
    if intake == "M":
        return dt.date(year, 5, 10)
    if intake == "S":
        return dt.date(year, 9, 10)
    return dt.date(year, 1, 10)


def class_days(start: dt.date, end: dt.date) -> List[dt.date]:
    out = []
    d = start
    while d <= end:
        if d.weekday() < 5:
            out.append(d)
        d += dt.timedelta(days=1)
    return out


def generate_attendance(students: pd.DataFrame, end_date: dt.date, seed: int) -> pd.DataFrame:
    rng = random.Random(seed)
    rows = []
    for _, r in students.iterrows():
        reg = str(r["REG. NO."])
        acc = r.get("ACC. NO.", "")
        regs = int(r.get("TOTAL REGISTRATIONS", 1) if not pd.isna(r.get("TOTAL REGISTRATIONS", 1)) else 1)
        parts = parse_reg(reg)
        if parts:
            current = semester_start(2000 + int(parts["yy"]), str(parts["intake"]))
        else:
            current = dt.date(2024, 1, 10)

        for sem in range(1, regs + 1):
            sem_start = current
            sem_end = min(sem_start + dt.timedelta(days=110), end_date)
            for d in class_days(sem_start, sem_end):
                status = rng.choices(
                    ["PRESENT", "PRESENT", "PRESENT", "PRESENT", "PRESENT", "LATE", "ABSENT"],
                    weights=[60, 10, 10, 10, 5, 3, 2],
                )[0]
                rows.append({
                    "REG_NO": reg,
                    "ACC_NO": acc,
                    "DATE": d,
                    "SEMESTER_INDEX": sem,
                    "STATUS": status,
                })
            current = sem_start + dt.timedelta(days=180)
            if current > end_date:
                break
    return pd.DataFrame(rows)


def build_dim_date(start: dt.date, end: dt.date) -> pd.DataFrame:
    rows = []
    d = start
    while d <= end:
        month = d.month
        if month in [1, 2, 3, 4]:
            acad_year = f"{d.year}/{d.year+1}"
            acad_term = "SEM1"
            intake_window = "January Cycle"
        elif month in [5, 6, 7, 8]:
            acad_year = f"{d.year}/{d.year+1}"
            acad_term = "MID-YEAR"
            intake_window = "May Cycle"
        else:
            acad_year = f"{d.year}/{d.year+1}"
            acad_term = "SEM2"
            intake_window = "September Cycle"
        rows.append({
            "date_key": int(d.strftime("%Y%m%d")),
            "full_date": d.isoformat(),
            "day": d.day,
            "day_name": d.strftime("%A"),
            "day_of_week_num": d.isoweekday(),
            "is_weekend": d.isoweekday() >= 6,
            "week_of_year": int(d.strftime("%V")),
            "month_num": d.month,
            "month_name": d.strftime("%B"),
            "month_short": d.strftime("%b"),
            "quarter": f"Q{((d.month - 1) // 3) + 1}",
            "year": d.year,
            "year_month": d.strftime("%Y-%m"),
            "academic_year": acad_year,
            "academic_term": acad_term,
            "intake_window": intake_window,
        })
        d += dt.timedelta(days=1)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Documentation
# ---------------------------------------------------------------------------

OVERALL_MD = """# University Analytics System
## Overall Project Documentation

This file summarizes the generated data assets, architecture, and analytics capabilities for the project.

- Student master data
- Course catalog
- Grades
- Transcript
- Academic performance
- Attendance
- Payments
- Sponsorships
- Calendar dimension
- BI dashboards
- Star schema architecture
- NextGen Query analytics
"""

UNIVERSITY_DOC_MD = """# University Analytics System
## Comprehensive Data Documentation, BI Dashboards, and Advanced Analytics

This project integrates academic, financial, and student engagement data to support institutional analytics.

Recommended dashboards:
1. Executive Overview
2. Academic Performance
3. Course Difficulty
4. Student Finance
5. Attendance Analytics
6. Student Risk Analytics

Advanced analytics:
- Attendance vs GPA
- Revenue by faculty
- Course failure analysis
- Program difficulty ranking
- Retake pressure
"""

ARCH_MD = """# University Analytics System
## Data Architecture Diagram, ERD, and Star Schema

Core dimensions:
- dim_student
- dim_program
- dim_faculty_department
- dim_course
- dim_date

Core facts:
- fact_grades
- fact_payments
- fact_sponsorships
- fact_attendance
- fact_academic_performance
- fact_transcript
"""

DATA_DESC_MD = """# UCU Analytics Demo Dataset
## data_description.md

This project dataset simulates a realistic university operating environment:
- students
- payments
- sponsorships
- grades
- attendance
- transcripts
- performance summaries
- date dimension
"""


# ---------------------------------------------------------------------------
# File writers
# ---------------------------------------------------------------------------

def save_csv_xlsx(df: pd.DataFrame, csv_path: Path, xlsx_path: Optional[Path] = None, sheet_name: str = "Main") -> None:
    df.to_csv(csv_path, index=False)
    if xlsx_path is not None:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)


def save_markdown(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------

@dataclass
class Inputs:
    source_students_15: Path
    source_students_16: Path
    faculties_departments: Optional[Path]
    fees_pdf: Optional[Path]


def generate_all(inputs: Inputs, output_dir: Path, target_rows: int = 5000) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load source student lists
    src15 = load_source_students(inputs.source_students_15)
    src16 = load_source_students(inputs.source_students_16)

    tuition_map = extract_tuition_map_from_pdf(inputs.fees_pdf) if inputs.fees_pdf else {}

    # 1) Students
    students15 = anonymize_students(src15, "L15", target_rows, tuition_map, src15, seed=20260304)
    students16 = anonymize_students(src16, "L16", target_rows, tuition_map, src16, seed=20260305)

    save_csv_xlsx(
        students15,
        output_dir / "students_list15_anonymized_5000_corrected_fee_logic.csv",
        output_dir / "students_list15_anonymized_5000_corrected_fee_logic.xlsx",
    )
    save_csv_xlsx(
        students16,
        output_dir / "students_list16_anonymized_5000_corrected_fee_logic.csv",
        output_dir / "students_list16_anonymized_5000_corrected_fee_logic.xlsx",
    )

    # 2) Course catalog
    programs = set(students15["PROGRAM"].dropna().astype(str).str.strip()).union(
        set(students16["PROGRAM"].dropna().astype(str).str.strip())
    )
    max_semesters = int(max(students15["TOTAL REGISTRATIONS"].fillna(1).max(), students16["TOTAL REGISTRATIONS"].fillna(1).max()))
    catalog = build_course_catalog(programs, max_semesters)
    save_csv_xlsx(
        catalog,
        output_dir / "course_catalog_ucu_actual_titles.csv",
        output_dir / "course_catalog_ucu_actual_titles.xlsx",
    )

    # 3) Grades
    grades15 = generate_grades(students15, catalog, "L15", seed=20260304)
    grades16 = generate_grades(students16, catalog, "L16", seed=20260305)
    save_csv_xlsx(
        grades15,
        output_dir / "student_grades_list15_updated_titles.csv",
        output_dir / "student_grades_list15_updated_titles.xlsx",
    )
    save_csv_xlsx(
        grades16,
        output_dir / "student_grades_list16_updated_titles.csv",
        output_dir / "student_grades_list16_updated_titles.xlsx",
    )

    # 4) Transcript
    transcript15 = build_transcript(grades15)
    transcript16 = build_transcript(grades16)
    save_csv_xlsx(
        transcript15,
        output_dir / "student_transcript_list15.csv",
        output_dir / "student_transcript_list15.xlsx",
        sheet_name="TRANSCRIPT",
    )
    save_csv_xlsx(
        transcript16,
        output_dir / "student_transcript_list16.csv",
        output_dir / "student_transcript_list16.xlsx",
        sheet_name="TRANSCRIPT",
    )

    # 5) Academic performance
    perf15 = build_academic_performance(grades15)
    perf16 = build_academic_performance(grades16)
    save_csv_xlsx(
        perf15,
        output_dir / "fact_student_academic_performance_list15.csv",
        output_dir / "fact_student_academic_performance_list15.xlsx",
    )
    save_csv_xlsx(
        perf16,
        output_dir / "fact_student_academic_performance_list16.csv",
        output_dir / "fact_student_academic_performance_list16.xlsx",
    )

    # 6) Payments
    payments15 = generate_payments(students15, "L15", seed=20260304)
    payments16 = generate_payments(students16, "L16", seed=20260305)
    save_csv_xlsx(
        payments15,
        output_dir / "student_payments_list15_realistic.csv",
        output_dir / "student_payments_list15_realistic.xlsx",
    )
    save_csv_xlsx(
        payments16,
        output_dir / "student_payments_list16_realistic.csv",
        output_dir / "student_payments_list16_realistic.xlsx",
    )

    # 7) Sponsorships
    sponsor15 = generate_sponsorships(students15, "L15", seed=20260304)
    sponsor16 = generate_sponsorships(students16, "L16", seed=20260305)
    save_csv_xlsx(
        sponsor15,
        output_dir / "student_sponsorships_list15.csv",
        output_dir / "student_sponsorships_list15.xlsx",
    )
    save_csv_xlsx(
        sponsor16,
        output_dir / "student_sponsorships_list16.csv",
        output_dir / "student_sponsorships_list16.xlsx",
    )

    # 8) Attendance (CSV primarily; Excel may exceed row limits)
    attendance_end = dt.date(2026, 3, 15)
    att15 = generate_attendance(students15, attendance_end, seed=20260304)
    att16 = generate_attendance(students16, attendance_end, seed=20260305)
    att15.to_csv(output_dir / "student_attendance_list15.csv", index=False)
    att16.to_csv(output_dir / "student_attendance_list16.csv", index=False)

    # 9) Academic progression
    progression15 = build_academic_progression(grades15, "L15")
    progression16 = build_academic_progression(grades16, "L16")
    save_csv_xlsx(
        progression15,
        output_dir / "academic_progression_list15.csv",
        output_dir / "academic_progression_list15.xlsx",
    )
    save_csv_xlsx(
        progression16,
        output_dir / "academic_progression_list16.csv",
        output_dir / "academic_progression_list16.xlsx",
    )

    # 10) Date dimension
    dim_date = build_dim_date(dt.date(2022, 1, 1), dt.date(2026, 12, 31))
    save_csv_xlsx(
        dim_date,
        output_dir / "dim_date_2022_2026.csv",
        output_dir / "dim_date_2022_2026.xlsx",
    )

    # 11) Documentation
    save_markdown(output_dir / "overall.md", OVERALL_MD)
    save_markdown(output_dir / "university_analytics_complete_documentation.md", UNIVERSITY_DOC_MD)
    save_markdown(output_dir / "data_architecture.md", ARCH_MD)
    save_markdown(output_dir / "data_description.md", DATA_DESC_MD)

    # Also preserve raw faculties_departments if supplied
    if inputs.faculties_departments and inputs.faculties_departments.exists():
        df_fd = pd.read_csv(inputs.faculties_departments)
        save_csv_xlsx(df_fd, output_dir / "faculties_departments.csv", output_dir / "faculties_departments.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Regenerate all synthetic datasets for the University Analytics System.")
    parser.add_argument("--source-students-15", type=Path, default=Path("students_list15.xlsx"), help="Path to source students_list15.xlsx")
    parser.add_argument("--source-students-16", type=Path, default=Path("students_list16.xlsx"), help="Path to source students_list16.xlsx")
    parser.add_argument("--faculties-departments", type=Path, default=Path("faculties_departments.csv"), help="Optional path to faculties_departments.csv")
    parser.add_argument("--fees-pdf", type=Path, default=Path("FEES STRUCTURE 2025 INTAKE.pdf"), help="Optional path to the fee structure PDF")
    parser.add_argument("--output-dir", type=Path, default=Path("generated_output"), help="Directory where all regenerated outputs will be written")
    parser.add_argument("--target-rows", type=int, default=5000, help="Target number of students per anonymized student list")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    inputs = Inputs(
        source_students_15=args.source_students_15,
        source_students_16=args.source_students_16,
        faculties_departments=args.faculties_departments if args.faculties_departments.exists() else None,
        fees_pdf=args.fees_pdf if args.fees_pdf.exists() else None,
    )
    generate_all(inputs, args.output_dir, target_rows=args.target_rows)
    print(f"Dataset regeneration complete. Outputs written to: {args.output_dir.resolve()}")


if __name__ == "__main__":
    main()
